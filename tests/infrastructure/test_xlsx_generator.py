"""Tests para T-053: XlsxGenerator."""
from __future__ import annotations

import io
from datetime import date, datetime

from openpyxl import load_workbook

from vulntrack.domain.value_objects.priority_score import PriorityBand
from vulntrack.domain.value_objects.severity import Severity
from vulntrack.infrastructure.reports.chart_builder import (
    EvolutionRow,
    PrioritizedFindingRow,
    ProjectRow,
    ReportData,
)
from vulntrack.infrastructure.reports.xlsx_generator import XlsxGenerator


def _sample_data() -> ReportData:
    return ReportData(
        period_label="Q2 2026",
        date_from=date(2026, 4, 1),
        date_to=date(2026, 6, 30),
        generated_at=datetime(2026, 6, 25, 10, 0, 0),
        author="Anderson Avendaño",
        total_vigentes=223,
        total_nuevas=118,
        total_tratadas=97,
        risk_score_portfolio=7.8,
        portfolio_metrics={
            Severity.CRITICAL: 12,
            Severity.HIGH: 45,
            Severity.MEDIUM: 98,
            Severity.LOW: 58,
            Severity.UNASSIGNED: 10,
        },
        new_portfolio_metrics={
            Severity.CRITICAL: 5,
            Severity.HIGH: 30,
            Severity.MEDIUM: 60,
            Severity.LOW: 23,
            Severity.UNASSIGNED: 0,
        },
        project_rows=[
            ProjectRow("daviplata-webview-frontend", 5, 10, 20, 15, 2, 52, 8.5),
            ProjectRow("backend-core", 2, 3, 8, 5, 0, 18, 4.2),
        ],
        evolution_rows=[
            EvolutionRow("daviplata-webview-frontend", 60, 52, -8, 8),
            EvolutionRow("backend-core", 14, 18, 4, 0),
        ],
        prioritized_findings=[
            PrioritizedFindingRow(
                vuln_id="CVE-2024-1234",
                component_name="log4j-core",
                component_version="2.14.0",
                project_name="daviplata-webview-frontend",
                severity=Severity.CRITICAL,
                cvss_v3_base_score=9.8,
                epss_score=0.85,
                is_kev=True,
                priority_score=92.5,
                priority_band=PriorityBand.CRITICAL,
            ),
        ],
        kev_hits=[],
    )


class TestXlsxGenerator:
    def setup_method(self) -> None:
        self.generator = XlsxGenerator()
        self.data = _sample_data()

    def test_generate_returns_bytes(self) -> None:
        result = self.generator.generate(self.data)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_generate_is_valid_xlsx(self) -> None:
        result = self.generator.generate(self.data)
        wb = load_workbook(io.BytesIO(result))
        assert wb is not None

    def test_sheet_names(self) -> None:
        result = self.generator.generate(self.data)
        wb = load_workbook(io.BytesIO(result))
        assert "Resumen" in wb.sheetnames
        assert "Estado" in wb.sheetnames
        assert "Nuevas" in wb.sheetnames
        assert "Evolución" in wb.sheetnames
        assert "Hallazgos Priorizados" in wb.sheetnames

    def test_resumen_kpis(self) -> None:
        result = self.generator.generate(self.data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Resumen"]
        all_values = [str(ws.cell(row=r, column=c).value or "") for r in range(1, 20) for c in range(1, 10)]
        assert "223" in all_values
        assert "118" in all_values

    def test_estado_project_data(self) -> None:
        result = self.generator.generate(self.data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Estado"]
        # Row 2 should be first project
        assert ws.cell(row=2, column=1).value == "daviplata-webview-frontend"
        assert ws.cell(row=2, column=2).value == 5  # critical
        assert ws.cell(row=2, column=7).value == 52  # total

    def test_estado_has_frozen_pane(self) -> None:
        result = self.generator.generate(self.data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Estado"]
        assert ws.freeze_panes == "A2"

    def test_evolucion_variacion_values(self) -> None:
        result = self.generator.generate(self.data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Evolución"]
        # First data row: daviplata, inicio=60, actual=52, variacion=-8, tratadas=8
        assert ws.cell(row=2, column=2).value == 60
        assert ws.cell(row=2, column=3).value == 52
        assert ws.cell(row=2, column=4).value == -8
        assert ws.cell(row=2, column=5).value == 8

    def test_prioritized_findings_sheet(self) -> None:
        result = self.generator.generate(self.data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Hallazgos Priorizados"]
        assert ws.cell(row=2, column=1).value == "CVE-2024-1234"
        assert ws.cell(row=2, column=6).value == 9.8   # cvss
        assert ws.cell(row=2, column=9).value == 92.5  # score

    def test_generate_with_empty_data(self) -> None:
        empty = ReportData(
            period_label="Q1 2026",
            date_from=date(2026, 1, 1),
            date_to=date(2026, 3, 31),
            generated_at=datetime(2026, 3, 31, 0, 0, 0),
            author="Test",
            total_vigentes=0,
            total_nuevas=0,
            total_tratadas=0,
            risk_score_portfolio=0.0,
        )
        result = self.generator.generate(empty)
        assert len(result) > 0
        wb = load_workbook(io.BytesIO(result))
        assert "Resumen" in wb.sheetnames
