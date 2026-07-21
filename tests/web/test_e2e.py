"""T-082: Pruebas end-to-end con SQLite en memoria y fixtures Q2 2026.

Setup: base de datos en memoria con proyectos, findings, snapshots y KEV
del período Q2 2026 (2026-04-01 a 2026-06-30).

KPIs de referencia:
  - Vigentes  : 223  (150 daviplata + 73 payments)
  - Nuevas    : 118  ( 80 daviplata +  38 payments, attributed_on en Q2)
  - Tratadas  :  97  ( 60 daviplata +  37 payments)
"""
from __future__ import annotations

import io
import os
import tempfile
from collections.abc import AsyncGenerator
from datetime import UTC, date, datetime, timedelta

import httpx
import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy.pool import NullPool
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from vulntrack.domain.entities.metric_snapshot import SnapshotSource
from vulntrack.domain.value_objects.severity import Severity
from vulntrack.infrastructure.persistence.database import Base
from vulntrack.infrastructure.persistence.orm_models import (
    FindingORM,
    KevEntryORM,
    MetricSnapshotORM,
    ProjectORM,
)

# ─── Constantes de fixtures ───────────────────────────────────────────────────

_UUID_DAVIPLATA = "11111111-1111-1111-1111-111111111111"
_UUID_PAYMENTS = "22222222-2222-2222-2222-222222222222"
_KEV_CVE = "CVE-2024-99001"

# Q2 2026: 2026-04-01 → 2026-06-30
_Q2_IN_RANGE = datetime(2026, 4, 15)   # naive — get_new_in_range compara sin tz
_Q2_OUT_RANGE = datetime(2026, 1, 15)  # antes del Q2, no cuenta como "nueva"

# ─── Helpers de generación ────────────────────────────────────────────────────


def _build_findings(
    project_uuid: str,
    total: int,
    new_in_q2: int,
    kev_cve: str | None,
    high_risk: bool,
    id_offset: int,
) -> list[FindingORM]:
    """Genera `total` findings activos; los primeros `new_in_q2` caen en Q2 2026."""
    now = datetime.now(UTC)
    rows: list[FindingORM] = []
    for i in range(total):
        in_q2 = i < new_in_q2
        is_kev_finding = kev_cve is not None and i == 0
        rows.append(
            FindingORM(
                project_uuid=project_uuid,
                dt_finding_uuid=f"{project_uuid[:8]}-f{id_offset + i:05d}",
                component_name=f"comp-{id_offset + i}",
                component_version="1.0.0",
                component_group=None,
                vuln_id=kev_cve if is_kev_finding else f"CVE-2024-{id_offset + i:05d}",
                vuln_source="NVD",
                severity=(
                    Severity.CRITICAL.value
                    if (i < 5 and high_risk)
                    else Severity.HIGH.value
                ),
                cvss_v3_base_score=9.8 if (i == 0 and high_risk) else (7.5 if high_risk else 5.0),
                epss_score=0.85 if (i == 0 and high_risk) else 0.10,
                epss_percentile=0.95 if (i == 0 and high_risk) else 0.50,
                attributed_on=_Q2_IN_RANGE if in_q2 else _Q2_OUT_RANGE,
                suppressed=False,
                last_synced_at=now,
                created_at=now,
                updated_at=now,
            )
        )
    return rows


async def _seed(factory: async_sessionmaker) -> None:  # type: ignore[type-arg]
    now = datetime.now(UTC)
    async with factory() as session:
        async with session.begin():
            # ── Proyectos ─────────────────────────────────────────────────────
            session.add_all(
                [
                    ProjectORM(
                        uuid=_UUID_DAVIPLATA,
                        name="daviplata-webview-frontend",
                        version="2.1.0",
                        description=None,
                        last_bom_import=now,
                        last_synced_at=now,
                        created_at=now,
                        updated_at=now,
                    ),
                    ProjectORM(
                        uuid=_UUID_PAYMENTS,
                        name="payments-service",
                        version="1.0.0",
                        description=None,
                        last_bom_import=now,
                        last_synced_at=now,
                        created_at=now,
                        updated_at=now,
                    ),
                ]
            )
            # ── Findings ──────────────────────────────────────────────────────
            # daviplata: 150 activos, 80 nuevos en Q2, 1 en KEV con CVSS=9.8/EPSS=0.85
            for row in _build_findings(_UUID_DAVIPLATA, 150, 80, _KEV_CVE, True, 1):
                session.add(row)
            # payments: 73 activos, 38 nuevos en Q2, sin KEV
            for row in _build_findings(_UUID_PAYMENTS, 73, 38, None, False, 10_000):
                session.add(row)

            # ── Snapshots ─────────────────────────────────────────────────────
            # daviplata: total_assigned inicio=200, actual=140 → tratadas=60
            #   total_assigned = critical+high+medium+low (no unassigned)
            session.add_all(
                [
                    MetricSnapshotORM(
                        project_uuid=_UUID_DAVIPLATA,
                        snapshot_date=date(2026, 3, 31),
                        critical=10, high=60, medium=80, low=50, unassigned=10,
                        total=210, risk_score=9.5,
                        source=SnapshotSource.DT_HISTORICAL.value,
                        created_at=now,
                    ),
                    MetricSnapshotORM(
                        project_uuid=_UUID_DAVIPLATA,
                        snapshot_date=date(2026, 6, 29),
                        critical=10, high=40, medium=70, low=20, unassigned=10,
                        total=150, risk_score=9.5,
                        source=SnapshotSource.DT_CURRENT.value,
                        created_at=now,
                    ),
                    # payments: total_assigned inicio=137, actual=100 → tratadas=37
                    MetricSnapshotORM(
                        project_uuid=_UUID_PAYMENTS,
                        snapshot_date=date(2026, 3, 31),
                        critical=5, high=30, medium=60, low=42, unassigned=5,
                        total=142, risk_score=4.2,
                        source=SnapshotSource.DT_HISTORICAL.value,
                        created_at=now,
                    ),
                    MetricSnapshotORM(
                        project_uuid=_UUID_PAYMENTS,
                        snapshot_date=date(2026, 6, 29),
                        critical=5, high=20, medium=55, low=20, unassigned=5,
                        total=105, risk_score=4.2,
                        source=SnapshotSource.DT_CURRENT.value,
                        created_at=now,
                    ),
                ]
            )
            # ── Catálogo KEV ──────────────────────────────────────────────────
            session.add(
                KevEntryORM(
                    cve_id=_KEV_CVE,
                    vendor_project="TestVendor",
                    product="WebComponent",
                    vulnerability_name="Critical KEV Test Vulnerability",
                    date_added=date(2024, 1, 15),
                    short_description="Exploited in the wild. Apply patch immediately.",
                    required_action="Apply vendor patch immediately.",
                    due_date=date(2024, 3, 15),
                    notes=None,
                    catalog_updated_at=now,
                )
            )


# ─── Fixture ──────────────────────────────────────────────────────────────────


@pytest.fixture
async def e2e_client() -> AsyncGenerator[AsyncClient, None]:
    from vulntrack.interfaces.web.dependencies import get_db
    from vulntrack.interfaces.web.main import create_app

    tmp_fd, db_path = tempfile.mkstemp(suffix=".db", prefix="vulntrack_e2e_")
    os.close(tmp_fd)
    db_url = f"sqlite+aiosqlite:///{db_path}"

    engine = create_async_engine(
        db_url,
        connect_args={"check_same_thread": False},
        poolclass=NullPool,
    )
    factory: async_sessionmaker[AsyncSession] = async_sessionmaker(
        engine, class_=AsyncSession, expire_on_commit=False
    )
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    await _seed(factory)

    async def _override_get_db() -> AsyncGenerator[AsyncSession, None]:
        async with factory() as session:
            try:
                yield session
                await session.commit()
            except Exception:
                await session.rollback()
                raise

    app = create_app()
    app.dependency_overrides[get_db] = _override_get_db

    # ASGITransport runs the ASGI app in the SAME event loop as the test.
    # No anyio blocking portal, no background threads → no Windows crash.
    async with AsyncClient(
        transport=ASGITransport(app=app, raise_app_exceptions=False),
        base_url="http://test",
    ) as client:
        yield client

    await engine.dispose()
    try:
        os.unlink(db_path)
    except OSError:
        pass


# ─── Tests ────────────────────────────────────────────────────────────────────


class TestE2EReportKpis:
    """T-082-a: Reporte .xlsx con KPIs exactos del Q2 2026."""

    async def test_xlsx_kpis_vigentes_nuevas_tratadas(self, e2e_client: AsyncClient) -> None:
        resp = await e2e_client.post(
            "/api/v1/reports/generate",
            json={"period": "quarterly", "quarter": "Q2", "year": 2026, "formats": ["xlsx"]},
        )
        assert resp.status_code == 200
        assert "spreadsheet" in resp.headers.get("content-type", "")

        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb["Resumen"]

        # Hoja Resumen: KPIs en filas 9-11, columna B
        assert ws["B9"].value == 223, f"vigentes esperadas=223, obtenidas={ws['B9'].value}"
        assert ws["B10"].value == 118, f"nuevas esperadas=118, obtenidas={ws['B10'].value}"
        assert ws["B11"].value == 97, f"tratadas esperadas=97, obtenidas={ws['B11'].value}"

    async def test_xlsx_content_disposition_filename(self, e2e_client: AsyncClient) -> None:
        resp = await e2e_client.post(
            "/api/v1/reports/generate",
            json={"period": "quarterly", "quarter": "Q2", "year": 2026, "formats": ["xlsx"]},
        )
        assert resp.status_code == 200
        cd = resp.headers.get("content-disposition", "")
        assert "Reporte_Portafolio_Q2_2026.xlsx" in cd


class TestE2EPrioritizedFindings:
    """T-082-b: Priorización — daviplata con KEV aparece primero."""

    async def test_kev_finding_ranks_first(self, e2e_client: AsyncClient) -> None:
        resp = await e2e_client.get("/api/v1/findings/prioritized")
        assert resp.status_code == 200
        page = resp.json()
        findings = page["items"]
        assert len(findings) > 0

        top = findings[0]
        assert top["is_kev"] is True
        assert top["vuln_id"] == _KEV_CVE

    async def test_kev_only_filter(self, e2e_client: AsyncClient) -> None:
        resp = await e2e_client.get("/api/v1/findings/prioritized", params={"kev_only": "true"})
        assert resp.status_code == 200
        page = resp.json()
        findings = page["items"]
        # Solo debe devolver el finding KEV
        assert all(f["is_kev"] for f in findings)
        assert len(findings) == 1
        assert findings[0]["vuln_id"] == _KEV_CVE


class TestE2ERemediationPlan:
    """T-082-c: Flujo completo de plan de remediación."""

    async def test_suggest_kev_task_target_date(self, e2e_client: AsyncClient) -> None:
        # Crear plan para daviplata
        resp = await e2e_client.post(
            f"/api/v1/remediation/plans?project_uuid={_UUID_DAVIPLATA}",
            json={"name": "Plan Q2 2026", "description": "Plan de remediación Q2"},
        )
        assert resp.status_code in (200, 201)
        plan_id = resp.json()["id"]

        # Generar sugerencias automáticas
        resp = await e2e_client.post(f"/api/v1/remediation/plans/{plan_id}/suggest")
        assert resp.status_code == 200
        tasks = resp.json()
        assert len(tasks) > 0

        # La tarea KEV debe tener target_date = hoy + 7 días
        kev_tasks = [t for t in tasks if _KEV_CVE in t.get("title", "")]
        assert len(kev_tasks) == 1, "Debe existir exactamente una tarea para el finding KEV"
        expected_date = (date.today() + timedelta(days=7)).isoformat()
        assert kev_tasks[0]["target_date"] == expected_date

    async def test_export_plan_xlsx(self, e2e_client: AsyncClient) -> None:
        # Crear plan para payments-service
        resp = await e2e_client.post(
            f"/api/v1/remediation/plans?project_uuid={_UUID_PAYMENTS}",
            json={"name": "Export Test Plan", "description": None},
        )
        assert resp.status_code in (200, 201)
        plan_id = resp.json()["id"]

        # Exportar como xlsx
        resp = await e2e_client.post(f"/api/v1/remediation/plans/{plan_id}/export?fmt=xlsx")
        assert resp.status_code == 200
        # xlsx es un ZIP → empieza con bytes "PK"
        assert resp.content[:2] == b"PK", "El xlsx debe ser un archivo ZIP válido"
        # Verificar que openpyxl puede abrirlo
        wb = load_workbook(io.BytesIO(resp.content))
        assert len(wb.sheetnames) > 0
