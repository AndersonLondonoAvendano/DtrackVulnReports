"""T-067: Caso de uso ExportPlan — exporta plan de remediación en .xlsx o .pdf.

Iteración 4: el plan materializa sus ítems como `TratamientoVulnerabilidad`
(no como el `RemediationTask` legado, retirado) -- se exportan sus tratamientos.
"""
from __future__ import annotations

from enum import StrEnum

from vulntrack.domain.entities.remediation import RemediationPlan
from vulntrack.domain.entities.vulnerability_treatment import TratamientoVulnerabilidad
from vulntrack.domain.exceptions import DomainError
from vulntrack.domain.ports.remediation_repository import RemediationRepository
from vulntrack.domain.ports.treatment_repository import TreatmentRepository


class ExportFormat(StrEnum):
    XLSX = "xlsx"
    PDF = "pdf"


class PlanNotFoundError(DomainError):
    def __init__(self, plan_id: int) -> None:
        super().__init__(f"Plan de remediación {plan_id} no encontrado")


class ExportPlanUseCase:
    def __init__(self, repo: RemediationRepository, treatment_repo: TreatmentRepository) -> None:
        self._repo = repo
        self._treatment_repo = treatment_repo

    async def execute(self, plan_id: int, fmt: ExportFormat) -> bytes:
        plan = await self._repo.get_plan(plan_id)
        if plan is None:
            raise PlanNotFoundError(plan_id)

        all_treatments = await self._treatment_repo.list_by_project(plan.project_uuid)
        treatments = [t for t in all_treatments if t.plan_id == plan_id]

        if fmt == ExportFormat.XLSX:
            return _export_xlsx(plan, treatments)
        return _export_pdf(plan, treatments)


def _export_xlsx(plan: RemediationPlan, treatments: list[TratamientoVulnerabilidad]) -> bytes:
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Plan de Remediación"  # type: ignore[union-attr]

    # Header
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = [
        "ID", "CVE / Vuln ID", "Componente", "Estado", "Prioridad", "Responsable",
        "Fecha objetivo", "Notas",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)  # type: ignore[union-attr]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Rows
    for row_idx, t in enumerate(treatments, start=2):
        ws.cell(row=row_idx, column=1, value=t.id)  # type: ignore[union-attr]
        ws.cell(row=row_idx, column=2, value=t.vuln_key)  # type: ignore[union-attr]
        ws.cell(row=row_idx, column=3, value=t.component_name or "")  # type: ignore[union-attr]
        ws.cell(row=row_idx, column=4, value=t.estado.value)  # type: ignore[union-attr]
        ws.cell(row=row_idx, column=5, value=t.priority_band.value)  # type: ignore[union-attr]
        ws.cell(row=row_idx, column=6, value=t.responsable or "")  # type: ignore[union-attr]
        ws.cell(row=row_idx, column=7, value=str(t.fecha_objetivo) if t.fecha_objetivo else "")  # type: ignore[union-attr]
        ws.cell(row=row_idx, column=8, value=t.notas or "")  # type: ignore[union-attr]

    # Autowidth
    for col_cells in ws.columns:  # type: ignore[union-attr]
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 45)  # type: ignore[union-attr]

    ws.freeze_panes = "A2"  # type: ignore[union-attr]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _export_pdf(plan: RemediationPlan, treatments: list[TratamientoVulnerabilidad]) -> bytes:
    rows_html = "".join(
        f"<tr><td>{t.id}</td><td>{t.vuln_key}</td><td>{t.component_name or ''}</td>"
        f"<td>{t.estado.value}</td><td>{t.priority_band.value}</td>"
        f"<td>{t.responsable or ''}</td><td>{t.fecha_objetivo or ''}</td></tr>"
        for t in treatments
    )
    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
@page {{ size: A4; margin: 2cm; }}
body {{ font-family: sans-serif; font-size: 10px; }}
h1 {{ color: #1F3864; font-size: 16px; }}
h2 {{ color: #1F3864; font-size: 12px; margin-top: 8px; }}
table {{ border-collapse: collapse; width: 100%; }}
th {{ background: #1F3864; color: white; padding: 4px 6px; text-align: left; font-size: 9px; }}
td {{ border: 1px solid #ccc; padding: 3px 5px; font-size: 9px; }}
</style>
</head>
<body>
<h1>Plan de Remediación: {plan.name}</h1>
<h2>Proyecto: {plan.project_uuid}</h2>
<p>Generado: {plan.updated_at.strftime('%Y-%m-%d %H:%M')}</p>
<table>
<tr>
<th>ID</th><th>CVE / Vuln ID</th><th>Componente</th><th>Estado</th>
<th>Prioridad</th><th>Responsable</th><th>Fecha objetivo</th>
</tr>
{rows_html}
</table>
</body></html>"""

    from weasyprint import HTML

    pdf = HTML(string=html).write_pdf()
    return pdf  # type: ignore[return-value]
