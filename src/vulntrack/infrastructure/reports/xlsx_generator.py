"""T-053: Generador de reportes Excel (.xlsx)."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from vulntrack.domain.value_objects.priority_score import PriorityBand
from vulntrack.domain.value_objects.severity import Severity
from vulntrack.infrastructure.reports.chart_builder import (
    ReportData,
)

# ─── Paleta ──────────────────────────────────────────────────────────────────

_FILLS: dict[Severity, PatternFill] = {
    Severity.CRITICAL: PatternFill("solid", fgColor="C00000"),
    Severity.HIGH: PatternFill("solid", fgColor="FF0000"),
    Severity.MEDIUM: PatternFill("solid", fgColor="FF9900"),
    Severity.LOW: PatternFill("solid", fgColor="FFFF00"),
    Severity.UNASSIGNED: PatternFill("solid", fgColor="D9D9D9"),
}

_FILL_NAVY = PatternFill("solid", fgColor="1F3864")
_FILL_GREEN = PatternFill("solid", fgColor="70AD47")
_FILL_RED = PatternFill("solid", fgColor="C00000")
_FILL_LIGHT_BLUE = PatternFill("solid", fgColor="DEEAF7")
_FILL_LIGHT_GREEN = PatternFill("solid", fgColor="CCFFCC")
_FILL_LIGHT_RED = PatternFill("solid", fgColor="FFCCCC")

_FONT_WHITE_BOLD = Font(color="FFFFFF", bold=True)
_FONT_NAVY_BOLD = Font(color="1F3864", bold=True)
_FONT_DARK_BOLD = Font(bold=True)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

_SEV_LABELS: dict[Severity, str] = {
    Severity.CRITICAL: "Crítica",
    Severity.HIGH: "Alta",
    Severity.MEDIUM: "Media",
    Severity.LOW: "Baja",
    Severity.UNASSIGNED: "Sin asignar",
}

_BAND_LABELS: dict[PriorityBand, str] = {
    PriorityBand.CRITICAL: "Inmediata",
    PriorityBand.HIGH: "Alta",
    PriorityBand.MEDIUM: "Media",
    PriorityBand.LOW: "Baja",
}

_BAND_FILLS: dict[PriorityBand, PatternFill] = {
    PriorityBand.CRITICAL: PatternFill("solid", fgColor="C00000"),
    PriorityBand.HIGH: PatternFill("solid", fgColor="FF9900"),
    PriorityBand.MEDIUM: PatternFill("solid", fgColor="FFFF00"),
    PriorityBand.LOW: PatternFill("solid", fgColor="D9D9D9"),
}


# ─── Helpers ─────────────────────────────────────────────────────────────────


def _header_cell(ws: object, row: int, col: int, value: str) -> None:  # type: ignore[type-arg]
    cell = ws.cell(row=row, column=col, value=value)  # type: ignore[attr-defined]
    cell.fill = _FILL_NAVY
    cell.font = _FONT_WHITE_BOLD
    cell.alignment = _ALIGN_CENTER


def _autowidth(ws: object) -> None:  # type: ignore[type-arg]
    """Ajusta el ancho de columna al contenido (heurística)."""
    for col in ws.columns:  # type: ignore[attr-defined]
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)  # type: ignore[attr-defined]


def _sev_cell(ws: object, row: int, col: int, sev: Severity, value: int) -> None:  # type: ignore[type-arg]
    cell = ws.cell(row=row, column=col, value=value)  # type: ignore[attr-defined]
    cell.alignment = _ALIGN_CENTER
    if value > 0:
        cell.fill = _FILLS[sev]
        cell.font = (
            _FONT_WHITE_BOLD
            if sev in (Severity.CRITICAL, Severity.HIGH)
            else _FONT_DARK_BOLD
        )


# ─── XlsxGenerator ───────────────────────────────────────────────────────────


class XlsxGenerator:
    """Implementa ReportGenerator para generar reportes Excel (.xlsx)."""

    def generate(self, data: ReportData) -> bytes:
        wb = Workbook()
        # Eliminar hoja por defecto
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)

        self._sheet_resumen(wb, data)
        self._sheet_estado(wb, data)
        self._sheet_nuevas(wb, data)
        self._sheet_evolucion(wb, data)
        self._sheet_prioritized(wb, data)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Hoja 1: Resumen ───────────────────────────────────────────────────────

    def _sheet_resumen(self, wb: Workbook, data: ReportData) -> None:
        ws = wb.create_sheet("Resumen")

        # Título
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = f"Reporte de Vulnerabilidades — {data.period_label}"
        title_cell.font = Font(bold=True, size=14, color="1F3864")
        title_cell.alignment = _ALIGN_CENTER

        # Metadatos
        meta = [
            ("Período:", data.period_label),
            ("Desde:", str(data.date_from)),
            ("Hasta:", str(data.date_to)),
            ("Generado:", data.generated_at.strftime("%Y-%m-%d %H:%M")),
            ("Elaborado por:", data.author),
        ]
        for i, (label, val) in enumerate(meta, start=2):
            ws.cell(row=i, column=1, value=label).font = _FONT_NAVY_BOLD
            ws.cell(row=i, column=2, value=val)

        # KPIs
        start_row = len(meta) + 3
        kpi_headers = ["KPI", "Valor"]
        for col, h in enumerate(kpi_headers, start=1):
            _header_cell(ws, start_row, col, h)

        kpis = [
            ("Vulnerabilidades vigentes", data.total_vigentes),
            ("Nuevas en el período", data.total_nuevas),
            ("Tratadas en el período", data.total_tratadas),
            ("Risk Score portafolio", round(data.risk_score_portfolio, 2)),
        ]
        for row_offset, (label, val) in enumerate(kpis, start=1):
            r = start_row + row_offset
            ws.cell(row=r, column=1, value=label).alignment = _ALIGN_LEFT
            val_cell = ws.cell(row=r, column=2, value=val)
            val_cell.alignment = _ALIGN_CENTER
            val_cell.font = Font(bold=True, size=13)
            if row_offset == 2:
                val_cell.font = Font(bold=True, size=13, color="C00000")
            elif row_offset == 3:
                val_cell.font = Font(bold=True, size=13, color="70AD47")

        # Distribución por severidad (vigentes)
        sev_start = start_row + len(kpis) + 2
        ws.cell(row=sev_start, column=1, value="Distribución por severidad (vigentes)").font = _FONT_NAVY_BOLD
        for col_off, sev in enumerate(Severity, start=1):
            _header_cell(ws, sev_start + 1, col_off, _SEV_LABELS[sev])
        for col_off, sev in enumerate(Severity, start=1):
            count = data.portfolio_metrics.get(sev, 0)
            _sev_cell(ws, sev_start + 2, col_off, sev, count)

        # Donut chart nativo
        chart = DoughnutChart()
        chart.title = "Vigentes por severidad"
        chart.style = 10
        labels_ref = Reference(ws, min_col=1, min_row=sev_start + 2, max_row=sev_start + 2 + len(list(Severity)) - 1)
        data_ref = Reference(ws, min_col=1, min_row=sev_start + 2, max_col=len(list(Severity)), max_row=sev_start + 2)
        chart.add_data(data_ref)
        chart.set_categories(Reference(ws, min_col=1, min_row=sev_start + 1, max_col=len(list(Severity)), max_row=sev_start + 1))
        ws.add_chart(chart, f"A{sev_start + 4}")

        _autowidth(ws)

    # ── Hoja 2: Estado (vigentes) ─────────────────────────────────────────────

    def _sheet_estado(self, wb: Workbook, data: ReportData) -> None:
        ws = wb.create_sheet("Estado")

        headers = ["Proyecto", "Crítica", "Alta", "Media", "Baja", "Sin asignar", "Total", "Risk Score"]
        for col, h in enumerate(headers, start=1):
            _header_cell(ws, 1, col, h)

        for row_idx, pr in enumerate(data.project_rows, start=2):
            ws.cell(row=row_idx, column=1, value=pr.name).alignment = _ALIGN_LEFT
            _sev_cell(ws, row_idx, 2, Severity.CRITICAL, pr.critical)
            _sev_cell(ws, row_idx, 3, Severity.HIGH, pr.high)
            _sev_cell(ws, row_idx, 4, Severity.MEDIUM, pr.medium)
            _sev_cell(ws, row_idx, 5, Severity.LOW, pr.low)
            _sev_cell(ws, row_idx, 6, Severity.UNASSIGNED, pr.unassigned)
            total_cell = ws.cell(row=row_idx, column=7, value=pr.total)
            total_cell.alignment = _ALIGN_CENTER
            total_cell.font = _FONT_DARK_BOLD
            ws.cell(row=row_idx, column=8, value=round(pr.risk_score, 2)).alignment = _ALIGN_CENTER

        ws.freeze_panes = "A2"
        _autowidth(ws)

        # Gráfico de barras nativo
        if data.project_rows:
            chart = BarChart()
            chart.type = "bar"
            chart.title = "Vulnerabilidades vigentes por proyecto"
            chart.y_axis.title = "Proyecto"
            chart.x_axis.title = "Cantidad"
            last_row = len(data.project_rows) + 1
            data_ref = Reference(ws, min_col=2, max_col=6, min_row=1, max_row=last_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            ws.add_chart(chart, f"J2")

    # ── Hoja 3: Nuevas ────────────────────────────────────────────────────────

    def _sheet_nuevas(self, wb: Workbook, data: ReportData) -> None:
        ws = wb.create_sheet("Nuevas")

        headers = ["Severidad", "Cantidad"]
        for col, h in enumerate(headers, start=1):
            _header_cell(ws, 1, col, h)

        sev_order = [
            Severity.CRITICAL, Severity.HIGH, Severity.MEDIUM,
            Severity.LOW, Severity.UNASSIGNED,
        ]
        for row_idx, sev in enumerate(sev_order, start=2):
            count = data.new_portfolio_metrics.get(sev, 0)
            ws.cell(row=row_idx, column=1, value=_SEV_LABELS[sev]).alignment = _ALIGN_CENTER
            _sev_cell(ws, row_idx, 2, sev, count)

        ws.freeze_panes = "A2"
        _autowidth(ws)

        # Donut nativo
        chart = DoughnutChart()
        chart.title = "Nuevas por severidad"
        chart.style = 10
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(sev_order) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=len(sev_order) + 1))
        ws.add_chart(chart, "D2")

    # ── Hoja 4: Evolución ─────────────────────────────────────────────────────

    def _sheet_evolucion(self, wb: Workbook, data: ReportData) -> None:
        ws = wb.create_sheet("Evolución")

        headers = ["Proyecto", "Inicio", "Actual", "Variación", "Tratadas"]
        for col, h in enumerate(headers, start=1):
            _header_cell(ws, 1, col, h)

        for row_idx, er in enumerate(data.evolution_rows, start=2):
            ws.cell(row=row_idx, column=1, value=er.name).alignment = _ALIGN_LEFT
            ws.cell(row=row_idx, column=2, value=er.inicio).alignment = _ALIGN_CENTER
            ws.cell(row=row_idx, column=3, value=er.actual).alignment = _ALIGN_CENTER

            var_cell = ws.cell(row=row_idx, column=4, value=er.variacion)
            var_cell.alignment = _ALIGN_CENTER
            var_cell.number_format = "+0;-0;0"
            if er.variacion > 0:
                var_cell.fill = _FILL_LIGHT_RED
            elif er.variacion < 0:
                var_cell.fill = _FILL_LIGHT_GREEN

            trat_cell = ws.cell(row=row_idx, column=5, value=er.tratadas)
            trat_cell.alignment = _ALIGN_CENTER
            if er.tratadas > 0:
                trat_cell.fill = _FILL_GREEN
                trat_cell.font = _FONT_WHITE_BOLD

        ws.freeze_panes = "A2"
        _autowidth(ws)

        # Gráfico agrupado inicio vs actual
        if data.evolution_rows:
            chart = BarChart()
            chart.title = "Inicio vs Actual"
            chart.style = 10
            last_row = len(data.evolution_rows) + 1
            data_ref = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=last_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "G2")

    # ── Hoja 5: Hallazgos priorizados ─────────────────────────────────────────

    def _sheet_prioritized(self, wb: Workbook, data: ReportData) -> None:
        ws = wb.create_sheet("Hallazgos Priorizados")

        headers = [
            "CVE / ID", "Componente", "Versión", "Proyecto",
            "Severidad", "CVSS", "EPSS", "KEV", "Score", "Prioridad",
        ]
        for col, h in enumerate(headers, start=1):
            _header_cell(ws, 1, col, h)

        for row_idx, fr in enumerate(data.prioritized_findings, start=2):
            ws.cell(row=row_idx, column=1, value=fr.vuln_id).alignment = _ALIGN_LEFT
            ws.cell(row=row_idx, column=2, value=fr.component_name).alignment = _ALIGN_LEFT
            ws.cell(row=row_idx, column=3, value=fr.component_version or "").alignment = _ALIGN_CENTER
            ws.cell(row=row_idx, column=4, value=fr.project_name).alignment = _ALIGN_LEFT

            sev_cell = ws.cell(row=row_idx, column=5, value=_SEV_LABELS.get(fr.severity, fr.severity))
            sev_cell.alignment = _ALIGN_CENTER
            sev_cell.fill = _FILLS[fr.severity]
            if fr.severity in (Severity.CRITICAL, Severity.HIGH):
                sev_cell.font = _FONT_WHITE_BOLD

            cvss_val = fr.cvss_v3_base_score if fr.cvss_v3_base_score is not None else ""
            ws.cell(row=row_idx, column=6, value=cvss_val).alignment = _ALIGN_CENTER

            epss_val = fr.epss_score if fr.epss_score is not None else ""
            ws.cell(row=row_idx, column=7, value=epss_val).alignment = _ALIGN_CENTER

            kev_cell = ws.cell(row=row_idx, column=8, value="Sí" if fr.is_kev else "No")
            kev_cell.alignment = _ALIGN_CENTER
            if fr.is_kev:
                kev_cell.fill = _FILL_RED
                kev_cell.font = _FONT_WHITE_BOLD

            score_cell = ws.cell(row=row_idx, column=9, value=round(fr.priority_score, 1))
            score_cell.alignment = _ALIGN_CENTER
            score_cell.font = Font(bold=True)

            band_cell = ws.cell(row=row_idx, column=10, value=_BAND_LABELS.get(fr.priority_band, fr.priority_band))
            band_cell.alignment = _ALIGN_CENTER
            band_cell.fill = _BAND_FILLS[fr.priority_band]
            if fr.priority_band == PriorityBand.CRITICAL:
                band_cell.font = _FONT_WHITE_BOLD

        ws.freeze_panes = "A2"
        _autowidth(ws)
